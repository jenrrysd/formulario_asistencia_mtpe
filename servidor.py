from flask import Flask, jsonify, request, send_from_directory, make_response
import json
import logging
import os
import socket
from datetime import datetime
import hashlib
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

# Configuración de seguridad
ADMIN_USUARIO = os.getenv('ADMIN_USUARIO', 'admin')  # Valor por defecto si no existe
ADMIN_CLAVE = os.getenv('ADMIN_CLAVE', 'usst2025')  # Valor por defecto si no existe
ADMIN_CLAVE_HASH = hashlib.sha256(ADMIN_CLAVE.encode()).hexdigest()

# Variable global para manejar sesión única
admin_sesion_activa = {
    "token": None,
    "timestamp": None,
    "ip": None
}


def obtener_ip_local():
    """Obtiene la IP LAN de la máquina (ej: 192.168.1.87)"""
    try:
        # Método 1: Conexión simulada a una IP externa (no se envía tráfico)
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))  # Google DNS
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        # Método 2: Fallback (puede devolver 127.0.0.1)
        return socket.gethostbyname(socket.gethostname())


# ✅ Desactivar logs HTTP innecesarios
log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

app = Flask(__name__)
ESTADO_ARCHIVO = 'estado.json'


def cargar_estado():
    if os.path.exists(ESTADO_ARCHIVO):
        with open(ESTADO_ARCHIVO, 'r', encoding='utf-8') as f:
            estado = json.load(f)
            # Calcular tiempo restante real basado en el tiempo transcurrido
            if estado.get('ultimo_inicio'):
                try:
                    inicio = datetime.fromisoformat(estado['ultimo_inicio'])
                    ahora = datetime.now()
                    transcurrido = int((ahora - inicio).total_seconds())
                    tiempo_inicial = estado.get('tiempo_inicial', 0)
                    tiempo_restante = max(0, tiempo_inicial - transcurrido)
                    estado['tiempo_restante'] = tiempo_restante
                except:
                    estado['tiempo_restante'] = 0
            return estado
    return {
        "tiempo_restante": 0,
        "tiempo_inicial": 0,
        "registros": [],
        "admin_logueado": False,
        "ultimo_inicio": None
    }


def guardar_estado(estado):
    with open(ESTADO_ARCHIVO, 'w', encoding='utf-8') as f:
        json.dump(estado, f, ensure_ascii=False, indent=2)


@app.route('/api/login', methods=['POST'])
def login_admin():
    data = request.json
    usuario = data.get('usuario', '').strip()
    clave = data.get('clave', '')
    cliente_ip = request.remote_addr

    # Validar credenciales
    if usuario == ADMIN_USUARIO:
        clave_hash = hashlib.sha256(clave.encode()).hexdigest()
        if clave_hash == ADMIN_CLAVE_HASH:
            # Verificar si ya hay una sesión activa
            if admin_sesion_activa["token"] is not None:
                # Hay una sesión activa, verificar si es de la misma IP
                if admin_sesion_activa["ip"] != cliente_ip:
                    return jsonify({
                        "ok": False,
                        "error": "Ya hay un administrador conectado desde otra ubicación. Solo se permite una sesión activa."
                    }), 403
                else:
                    # Misma IP, permitir reconexión (renovar token)
                    pass

            # Generar nuevo token único
            nuevo_token = hashlib.sha256(
                f"{usuario}{clave}{datetime.now().isoformat()}{cliente_ip}".encode()).hexdigest()[:16]

            # Actualizar sesión activa
            admin_sesion_activa["token"] = nuevo_token
            admin_sesion_activa["timestamp"] = datetime.now().isoformat()
            admin_sesion_activa["ip"] = cliente_ip

            return jsonify({
                "ok": True,
                "mensaje": "Login exitoso",
                "admin_token": nuevo_token
            })

    return jsonify({
        "ok": False,
        "error": "Credenciales incorrectas"
    }), 401


def verificar_token_admin(token, cliente_ip):
    """Verifica si el token es válido y corresponde a la sesión activa"""
    if not token or admin_sesion_activa["token"] is None:
        return False

    # Verificar token y IP
    if (admin_sesion_activa["token"] == token and
            admin_sesion_activa["ip"] == cliente_ip):
        return True

    return False


@app.route('/api/verificar-sesion', methods=['POST'])
def verificar_sesion():
    """Endpoint para verificar si la sesión sigue siendo válida"""
    data = request.json
    token = data.get('token', '')
    cliente_ip = request.remote_addr

    if verificar_token_admin(token, cliente_ip):
        return jsonify({
            "ok": True,
            "sesion_valida": True,
            "mensaje": "Sesión válida"
        })
    else:
        return jsonify({
            "ok": False,
            "sesion_valida": False,
            "mensaje": "Sesión inválida o expirada"
        }), 401


@app.route('/api/cerrar-sesion', methods=['POST'])
def cerrar_sesion():
    """Endpoint para cerrar la sesión activa"""
    data = request.json
    token = data.get('token', '')
    cliente_ip = request.remote_addr

    # Verificar que quien cierra sea el usuario logueado
    if verificar_token_admin(token, cliente_ip):
        # Limpiar sesión activa
        admin_sesion_activa["token"] = None
        admin_sesion_activa["timestamp"] = None
        admin_sesion_activa["ip"] = None

        return jsonify({
            "ok": True,
            "mensaje": "Sesión cerrada correctamente"
        })
    else:
        return jsonify({
            "ok": False,
            "mensaje": "No tienes una sesión activa válida"
        }), 401


@app.route('/api/estado', methods=['GET'])
def obtener_estado():
    estado = cargar_estado()
    # Agregar información de sesión (sin exponer datos sensibles)
    estado["sesion_admin_activa"] = admin_sesion_activa["token"] is not None
    return jsonify(estado)


@app.route('/api/iniciar', methods=['POST'])
def iniciar_temporizador():
    # Verificar token de administrador
    data = request.json
    token = data.get('admin_token', '')
    if not verificar_token_admin(token, request.remote_addr):
        return jsonify({"ok": False, "error": "Token de administrador inválido"}), 401

    minutos = data.get('minutos', 30)
    tiempo_segundos = minutos * 60

    estado = cargar_estado()
    estado['tiempo_restante'] = tiempo_segundos
    # Guardar tiempo inicial para cálculos
    estado['tiempo_inicial'] = tiempo_segundos
    estado['ultimo_inicio'] = datetime.now().isoformat()
    guardar_estado(estado)

    return jsonify({
        "ok": True,
        "tiempo": estado['tiempo_restante'],
        "minutos": minutos,
        "mensaje": f"Temporizador iniciado por {minutos} minutos"
    })


@app.route('/api/registrar', methods=['POST'])
def registrar_asistencia():
    data = request.json
    estado = cargar_estado()

    # Verificar si el tiempo ha expirado
    if estado['tiempo_restante'] <= 0:
        return jsonify({
            "ok": False,
            "error": "El tiempo para registrar asistencias ha expirado"
        }), 400

    # Añadir fecha/hora del servidor (más confiable)
    data['fecha_hora_servidor'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    estado['registros'].append(data)
    guardar_estado(estado)

    return jsonify({
        "ok": True,
        "total": len(estado['registros']),
        "tiempo_restante": estado['tiempo_restante']
    })


@app.route('/api/extender', methods=['POST'])
def extender_temporizador():
    # Verificar token de administrador
    data = request.json
    token = data.get('admin_token', '')
    if not verificar_token_admin(token, request.remote_addr):
        return jsonify({"ok": False, "error": "Token de administrador inválido"}), 401

    minutos_extra = data.get('minutos', 10)

    estado = cargar_estado()

    # Si hay un temporizador activo, extender el tiempo
    if estado.get('ultimo_inicio') and estado.get('tiempo_inicial', 0) > 0:
        # Calcular el nuevo tiempo inicial total
        segundos_extra = minutos_extra * 60
        estado['tiempo_inicial'] += segundos_extra

        # Recalcular el tiempo restante
        try:
            inicio = datetime.fromisoformat(estado['ultimo_inicio'])
            ahora = datetime.now()
            transcurrido = int((ahora - inicio).total_seconds())
            tiempo_restante = max(0, estado['tiempo_inicial'] - transcurrido)
            estado['tiempo_restante'] = tiempo_restante
        except:
            estado['tiempo_restante'] = segundos_extra

        guardar_estado(estado)

        return jsonify({
            "ok": True,
            "tiempo_restante": estado['tiempo_restante'],
            "minutos_agregados": minutos_extra,
            "mensaje": f"Se agregaron {minutos_extra} minutos al temporizador"
        })
    else:
        return jsonify({
            "ok": False,
            "error": "No hay un temporizador activo para extender"
        }), 400


@app.route('/api/detener', methods=['POST'])
def detener_temporizador():
    # Verificar token de administrador
    data = request.json
    token = data.get('admin_token', '')
    if not verificar_token_admin(token, request.remote_addr):
        return jsonify({"ok": False, "error": "Token de administrador inválido"}), 401

    estado = cargar_estado()
    estado['tiempo_restante'] = 0
    estado['tiempo_inicial'] = 0
    estado['ultimo_inicio'] = None
    guardar_estado(estado)

    return jsonify({
        "ok": True,
        "mensaje": "Temporizador detenido y formulario cerrado"
    })


@app.route('/api/limpiar', methods=['POST'])
def limpiar_registros():
    # Verificar token de administrador
    data = request.json
    token = data.get('admin_token', '')
    if not verificar_token_admin(token, request.remote_addr):
        return jsonify({"ok": False, "error": "Token de administrador inválido"}), 401

    estado = cargar_estado()
    estado['registros'] = []
    guardar_estado(estado)
    return jsonify({"ok": True, "total": 0})


@app.route('/api/descargar-excel', methods=['GET'])
def descargar_excel():
    estado = cargar_estado()
    if not estado['registros']:
        return 'No hay registros', 404

    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E56A0",
                              end_color="1E56A0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = ['APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRES', 'DNI', 'NÚMERO TELEFÓNICO',
               'CORREO ELECTRÓNICO', 'ÁREA ORGANIZACIONAL', 'CENTRO DE TRABAJO', 'FECHA/HORA DE ASISTENCIA']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Datos
    for row, registro in enumerate(estado['registros'], 2):
        datos = [
            registro.get('apellido_paterno', ''),
            registro.get('apellido_materno', ''),
            registro.get('nombres', ''),
            registro.get('dni', ''),
            registro.get('cargo_numero_telefonico', ''),
            registro.get('grupo_correo_electronico', ''),
            registro.get('area_organizacional', ''),
            registro.get('centro_trabajo', ''),
            registro.get('fecha_hora_servidor', '')
        ]

        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Ajustar ancho de columnas
    column_widths = [22, 22, 22, 12, 25, 30, 25, 25, 24]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(
            row=1, column=col).column_letter].width = width

    # Agregar información adicional
    info_row = len(estado['registros']) + 3
    ws.cell(row=info_row, column=1,
            value="Total de registros:").font = Font(bold=True)
    cell_total = ws.cell(row=info_row, column=2,
                         value=len(estado['registros']))
    cell_total.alignment = Alignment(horizontal="left")

    ws.cell(row=info_row + 1, column=1,
            value="Fecha de generación:").font = Font(bold=True)
    cell_fecha = ws.cell(row=info_row + 1, column=2,
                         value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
    cell_fecha.alignment = Alignment(horizontal="left")

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers[
        'Content-Disposition'] = f'attachment; filename="asistencias_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    return response

# Servir archivos estáticos (HTML, CSS, JS)


@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

# ✅ Sirve CSS, JS y otros archivos desde la misma carpeta


@app.route('/<path:filename>')
def archivos_estaticos(filename):
    # Solo permitir archivos que existan y no sean peligrosos
    if os.path.isfile(filename) and not filename.startswith('.'):
        return send_from_directory('.', filename)
    return "Archivo no encontrado", 404


if __name__ == '__main__':
    ip_local = obtener_ip_local()
    print(f"Servidor iniciado en http://{ip_local}:8080")
    print("Servidor iniciado en http://localhost:8080")
    print("Presiona Ctrl+C para detener el servidor.")

    app.run(host='0.0.0.0', port=8080, debug=True)
